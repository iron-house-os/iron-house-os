from io import BytesIO

from openpyxl import load_workbook

from app.schemas.estimate import (
    EstimateCreate,
    EstimateLineItem,
    EstimateMarkup,
    EstimateRiskAllowance,
    EstimateUnit,
    VendorQuoteInput,
)
from app.services.estimate_workbooks import build_estimate_workbook, workbook_filename

EXPECTED_SHEETS = [
    "Summary",
    "Line Items",
    "Production Rates",
    "Markups",
    "Risks",
    "Assumptions",
    "Exclusions",
    "Quote Comparison",
]


def test_workbook_contains_complete_estimator_sheet_contract() -> None:
    payload = EstimateCreate(
        project_name="Marine Drive Parking Lot",
        project_code="WR26-012",
        line_items=[
            EstimateLineItem(
                code="PIPE-001",
                description="Supply and install PVC storm pipe",
                quantity=100,
                unit=EstimateUnit.metre,
                direct_unit_cost=50,
                vendor_quotes=[
                    VendorQuoteInput(
                        supplier="EMCO",
                        scope="PVC pipe supply",
                        amount=4200,
                        is_selected=True,
                        selection_reason="Complete scope and confirmed delivery",
                        qualification_notes=["Scope reviewed"],
                        notes="Qualified selected quote",
                    )
                ],
            )
        ],
        risks=[
            EstimateRiskAllowance(
                description="Unknown utilities",
                amount=10000,
                probability=0.25,
            )
        ],
        markup=EstimateMarkup(
            contingency_percent=5,
            overhead_percent=10,
            profit_percent=15,
        ),
        assumptions=["Normal working hours"],
        exclusions=["Contaminated soil disposal"],
    )

    workbook = load_workbook(BytesIO(build_estimate_workbook(payload)), data_only=False)

    assert workbook.sheetnames == EXPECTED_SHEETS
    assert workbook["Assumptions"]["A4"].value == "Normal working hours"
    assert workbook["Exclusions"]["A4"].value == "Contaminated soil disposal"
    assert workbook["Line Items"]["L6"].value == "=SUM(L4:L4)"
    assert workbook["Quote Comparison"]["B4"].value == "EMCO"
    assert workbook["Quote Comparison"]["E4"].value == "Yes"
    assert workbook["Quote Comparison"]["F4"].value == "Yes"
    assert workbook["Quote Comparison"]["G4"].value == "Complete scope and confirmed delivery"
    assert workbook["Quote Comparison"]["H4"].value == "Scope reviewed"


def test_workbook_records_empty_assumptions_exclusions_and_line_items() -> None:
    workbook = load_workbook(
        BytesIO(build_estimate_workbook(EstimateCreate(project_name="Empty estimate"))),
        data_only=False,
    )

    assert workbook["Assumptions"]["A4"].value == "No assumptions entered."
    assert workbook["Exclusions"]["A4"].value == "No exclusions entered."
    assert workbook["Line Items"]["L5"].value == "=0"


def test_workbook_filename_is_safe_and_bounded() -> None:
    filename = workbook_filename("Marine Drive / Parking Lot", "WR26-012")

    assert filename == "WR26-012_Marine_Drive___Parking_Lot_Estimate.xlsx"
    assert len(filename) <= 120
