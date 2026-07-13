from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.schemas.estimate import EstimateCreate, EstimateSummary
from app.services.estimates import calculate_estimate, get_rate_library

TITLE_FILL = "1F4E79"
HEADER_FILL = "D9EAF7"
SUBTOTAL_FILL = "E2F0D9"
WARNING_FILL = "FCE4D6"
WHITE = "FFFFFF"
MONEY_FORMAT = '$#,##0.00;[Red]-$#,##0.00'
PERCENT_FORMAT = "0.00%"
NUMBER_FORMAT = "#,##0.00"


def build_estimate_workbook(payload: EstimateCreate) -> bytes:
    summary = calculate_estimate(payload)
    workbook = Workbook()
    workbook.remove(workbook.active)

    _build_summary_sheet(workbook, summary)
    _build_line_items_sheet(workbook, summary)
    _build_production_rates_sheet(workbook)
    _build_markups_sheet(workbook, payload, summary)
    _build_risks_sheet(workbook, payload)
    _build_assumptions_sheet(workbook, summary)
    _build_exclusions_sheet(workbook, summary)
    _build_quote_comparison_sheet(workbook, payload)

    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def _build_summary_sheet(workbook: Workbook, summary: EstimateSummary) -> None:
    sheet = workbook.create_sheet("Summary")
    _title(sheet, "Iron House Estimate Summary")
    rows: list[tuple[str, Any, str | None]] = [
        ("Project", summary.project_name, None),
        ("Project Code", summary.project_code or "", None),
        ("Direct Cost", summary.direct_cost, MONEY_FORMAT),
        ("Indirect Cost", summary.indirect_cost, MONEY_FORMAT),
        ("Risk Allowance", summary.risk_cost, MONEY_FORMAT),
        ("Subtotal Before Markup", summary.subtotal_before_markup, MONEY_FORMAT),
        ("Contingency", summary.contingency, MONEY_FORMAT),
        ("Bonding", summary.bonding, MONEY_FORMAT),
        ("Insurance", summary.insurance, MONEY_FORMAT),
        ("Overhead", summary.overhead, MONEY_FORMAT),
        ("Profit", summary.profit, MONEY_FORMAT),
        ("Final Bid Price", summary.final_price, MONEY_FORMAT),
        ("Gross Margin", summary.gross_margin_percent / 100, PERCENT_FORMAT),
    ]

    for row_index, (label, value, number_format) in enumerate(rows, start=3):
        sheet.cell(row=row_index, column=1, value=label)
        value_cell = sheet.cell(row=row_index, column=2, value=value)
        if number_format:
            value_cell.number_format = number_format
        if label in {"Subtotal Before Markup", "Final Bid Price"}:
            _fill_row(sheet, row_index, 1, 2, SUBTOTAL_FILL)
            sheet.cell(row=row_index, column=1).font = Font(bold=True)
            value_cell.font = Font(bold=True)

    sheet.cell(row=18, column=1, value="Category")
    sheet.cell(row=18, column=2, value="Amount")
    _header_row(sheet, 18, 1, 2)
    categories = summary.category_breakdown.model_dump()
    for row_index, (category, amount) in enumerate(categories.items(), start=19):
        sheet.cell(row=row_index, column=1, value=category.replace("_", " ").title())
        sheet.cell(row=row_index, column=2, value=amount).number_format = MONEY_FORMAT

    _apply_borders(sheet, 3, 1, 25, 2)
    _set_widths(sheet, {"A": 26, "B": 18})


def _build_line_items_sheet(workbook: Workbook, summary: EstimateSummary) -> None:
    sheet = workbook.create_sheet("Line Items")
    _title(sheet, "Estimate Line Items")
    headers = [
        "Code",
        "Description",
        "Type",
        "Qty",
        "Unit",
        "Hours",
        "Labour",
        "Equipment",
        "Material",
        "Disposal",
        "Subcontract",
        "Direct Cost",
        "Unit Cost",
        "Selected Supplier",
    ]
    _write_headers(sheet, headers, row=3)

    for row_index, item in enumerate(summary.line_items, start=4):
        values = [
            item.code or "",
            item.description,
            item.item_type.value,
            item.quantity,
            item.unit.value,
            item.hours,
            item.labour_cost,
            item.equipment_cost,
            item.material_cost,
            item.disposal_cost,
            item.subcontract_cost,
            item.direct_cost,
            item.unit_cost,
            item.selected_quote_supplier or "",
        ]
        _write_row(sheet, row_index, values)
        for column in range(7, 14):
            sheet.cell(row=row_index, column=column).number_format = MONEY_FORMAT
        sheet.cell(row=row_index, column=4).number_format = NUMBER_FORMAT
        sheet.cell(row=row_index, column=6).number_format = NUMBER_FORMAT

    total_row = len(summary.line_items) + 5
    sheet.cell(row=total_row, column=11, value="Total Direct Cost")
    total_formula = f"=SUM(L4:L{total_row - 2})" if summary.line_items else "=0"
    sheet.cell(row=total_row, column=12, value=total_formula)
    sheet.cell(row=total_row, column=12).number_format = MONEY_FORMAT
    _fill_row(sheet, total_row, 11, 12, SUBTOTAL_FILL)
    sheet.cell(row=total_row, column=11).font = Font(bold=True)
    sheet.cell(row=total_row, column=12).font = Font(bold=True)

    _apply_borders(sheet, 3, 1, total_row, len(headers))
    _set_widths(
        sheet,
        {
            "A": 12,
            "B": 34,
            "C": 16,
            "D": 10,
            "E": 10,
            "F": 10,
            "G": 14,
            "H": 14,
            "I": 14,
            "J": 14,
            "K": 14,
            "L": 16,
            "M": 14,
            "N": 22,
        },
    )


def _build_production_rates_sheet(workbook: Workbook) -> None:
    sheet = workbook.create_sheet("Production Rates")
    _title(sheet, "Default Production Rate Library")
    headers = ["Activity", "Description", "Unit", "Production / Hr", "Crew", "Equipment", "Notes"]
    _write_headers(sheet, headers, row=3)

    for row_index, rate in enumerate(get_rate_library().production_rates, start=4):
        crew = "; ".join(
            f"{member.quantity:g}x {member.role} @ ${member.hourly_rate:g}/hr"
            for member in rate.crew
        )
        equipment = "; ".join(
            f"{resource.quantity:g}x {resource.name} @ ${resource.hourly_rate:g}/hr"
            for resource in rate.equipment
        )
        _write_row(
            sheet,
            row_index,
            [
                rate.activity.value,
                rate.description,
                rate.unit.value,
                rate.production_rate_per_hour,
                crew,
                equipment,
                rate.notes or "",
            ],
        )
        sheet.cell(row=row_index, column=4).number_format = NUMBER_FORMAT

    _apply_borders(sheet, 3, 1, len(get_rate_library().production_rates) + 3, len(headers))
    _set_widths(sheet, {"A": 24, "B": 42, "C": 10, "D": 16, "E": 44, "F": 44, "G": 42})


def _build_markups_sheet(workbook: Workbook, payload: EstimateCreate, summary: EstimateSummary) -> None:
    sheet = workbook.create_sheet("Markups")
    _title(sheet, "Markup and Bid Build-Up")
    rows = [
        ("Contingency %", payload.markup.contingency_percent / 100, PERCENT_FORMAT),
        ("Bonding %", payload.markup.bonding_percent / 100, PERCENT_FORMAT),
        ("Insurance %", payload.markup.insurance_percent / 100, PERCENT_FORMAT),
        ("Overhead %", payload.markup.overhead_percent / 100, PERCENT_FORMAT),
        ("Profit %", payload.markup.profit_percent / 100, PERCENT_FORMAT),
        ("Subtotal Before Markup", summary.subtotal_before_markup, MONEY_FORMAT),
        ("Contingency", summary.contingency, MONEY_FORMAT),
        ("Bonding", summary.bonding, MONEY_FORMAT),
        ("Insurance", summary.insurance, MONEY_FORMAT),
        ("Overhead", summary.overhead, MONEY_FORMAT),
        ("Profit", summary.profit, MONEY_FORMAT),
        ("Final Bid Price", summary.final_price, MONEY_FORMAT),
    ]
    for row_index, (label, value, number_format) in enumerate(rows, start=3):
        sheet.cell(row=row_index, column=1, value=label)
        sheet.cell(row=row_index, column=2, value=value).number_format = number_format
        if label == "Final Bid Price":
            _fill_row(sheet, row_index, 1, 2, SUBTOTAL_FILL)
            sheet.cell(row=row_index, column=1).font = Font(bold=True)
            sheet.cell(row=row_index, column=2).font = Font(bold=True)
    _apply_borders(sheet, 3, 1, len(rows) + 2, 2)
    _set_widths(sheet, {"A": 28, "B": 18})


def _build_risks_sheet(workbook: Workbook, payload: EstimateCreate) -> None:
    sheet = workbook.create_sheet("Risks")
    _title(sheet, "Risk Allowances")
    headers = ["Description", "Amount", "Probability", "Expected Cost", "Notes"]
    _write_headers(sheet, headers, row=3)
    if not payload.risks:
        sheet.cell(row=4, column=1, value="No risk allowances entered.")
    for row_index, risk in enumerate(payload.risks, start=4):
        probability = risk.probability if risk.probability is not None else 1
        _write_row(
            sheet,
            row_index,
            [risk.description, risk.amount, probability, risk.amount * probability, risk.notes or ""],
        )
        sheet.cell(row=row_index, column=2).number_format = MONEY_FORMAT
        sheet.cell(row=row_index, column=3).number_format = PERCENT_FORMAT
        sheet.cell(row=row_index, column=4).number_format = MONEY_FORMAT
    _apply_borders(sheet, 3, 1, max(4, len(payload.risks) + 3), len(headers))
    _set_widths(sheet, {"A": 34, "B": 16, "C": 14, "D": 16, "E": 42})


def _build_assumptions_sheet(workbook: Workbook, summary: EstimateSummary) -> None:
    sheet = workbook.create_sheet("Assumptions")
    _title(sheet, "Estimate Assumptions")
    sheet.cell(row=3, column=1, value="Assumption")
    _header_row(sheet, 3, 1, 1)
    if not summary.assumptions:
        sheet.cell(row=4, column=1, value="No assumptions entered.")
    for row_index, assumption in enumerate(summary.assumptions, start=4):
        sheet.cell(row=row_index, column=1, value=assumption)
    _apply_borders(sheet, 3, 1, max(4, len(summary.assumptions) + 3), 1)
    _set_widths(sheet, {"A": 100})


def _build_exclusions_sheet(workbook: Workbook, summary: EstimateSummary) -> None:
    sheet = workbook.create_sheet("Exclusions")
    _title(sheet, "Estimate Exclusions")
    sheet.cell(row=3, column=1, value="Exclusion")
    _header_row(sheet, 3, 1, 1)
    if not summary.exclusions:
        sheet.cell(row=4, column=1, value="No exclusions entered.")
    for row_index, exclusion in enumerate(summary.exclusions, start=4):
        sheet.cell(row=row_index, column=1, value=exclusion)
    _apply_borders(sheet, 3, 1, max(4, len(summary.exclusions) + 3), 1)
    _set_widths(sheet, {"A": 100})

def _build_quote_comparison_sheet(workbook: Workbook, payload: EstimateCreate) -> None:
    sheet = workbook.create_sheet("Quote Comparison")
    _title(sheet, "Supplier and Subcontractor Quote Comparison")
    headers = [
        "Line Item",
        "Supplier",
        "Scope",
        "Amount",
        "Qualified",
        "Selected",
        "Selection Reason",
        "Qualification Notes",
        "Notes",
    ]
    _write_headers(sheet, headers, row=3)
    row_index = 4
    for item in payload.line_items:
        for quote in item.vendor_quotes:
            _write_row(
                sheet,
                row_index,
                [
                    item.description,
                    quote.supplier,
                    quote.scope,
                    quote.amount,
                    "Yes" if quote.is_qualified else "No",
                    "Yes" if quote.is_selected else "No",
                    quote.selection_reason or "",
                    "; ".join(quote.qualification_notes),
                    quote.notes or "",
                ],
            )
            sheet.cell(row=row_index, column=4).number_format = MONEY_FORMAT
            if quote.is_selected:
                _fill_row(sheet, row_index, 1, len(headers), SUBTOTAL_FILL)
            elif not quote.is_qualified:
                _fill_row(sheet, row_index, 1, len(headers), WARNING_FILL)
            row_index += 1
    if row_index == 4:
        sheet.cell(row=4, column=1, value="No vendor quotes entered.")
    _apply_borders(sheet, 3, 1, max(4, row_index - 1), len(headers))
    _set_widths(
        sheet,
        {
            "A": 34,
            "B": 26,
            "C": 34,
            "D": 16,
            "E": 12,
            "F": 12,
            "G": 42,
            "H": 42,
            "I": 42,
        },
    )


def _title(sheet: Worksheet, title: str) -> None:
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    cell = sheet.cell(row=1, column=1, value=title)
    cell.font = Font(bold=True, size=16, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=TITLE_FILL)
    cell.alignment = Alignment(horizontal="left")


def _write_headers(sheet: Worksheet, headers: list[str], row: int) -> None:
    for column, header in enumerate(headers, start=1):
        sheet.cell(row=row, column=column, value=header)
    _header_row(sheet, row, 1, len(headers))


def _write_row(sheet: Worksheet, row: int, values: list[Any]) -> None:
    for column, value in enumerate(values, start=1):
        sheet.cell(row=row, column=column, value=value)


def _header_row(sheet: Worksheet, row: int, start_column: int, end_column: int) -> None:
    for column in range(start_column, end_column + 1):
        cell = sheet.cell(row=row, column=column)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _fill_row(sheet: Worksheet, row: int, start_column: int, end_column: int, fill_color: str) -> None:
    for column in range(start_column, end_column + 1):
        sheet.cell(row=row, column=column).fill = PatternFill("solid", fgColor=fill_color)


def _apply_borders(sheet: Worksheet, start_row: int, start_col: int, end_row: int, end_col: int) -> None:
    side = Side(style="thin", color="D9D9D9")
    border = Border(left=side, right=side, top=side, bottom=side)
    for row in range(start_row, end_row + 1):
        for column in range(start_col, end_col + 1):
            cell = sheet.cell(row=row, column=column)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _set_widths(sheet: Worksheet, widths: dict[str, int]) -> None:
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    for row in range(1, sheet.max_row + 1):
        sheet.row_dimensions[row].height = 20
    sheet.freeze_panes = "A4"


def workbook_filename(project_name: str, project_code: str | None = None) -> str:
    raw_name = f"{project_code + '_' if project_code else ''}{project_name}_Estimate.xlsx"
    safe_name = "".join(character if character.isalnum() or character in "-_." else "_" for character in raw_name)
    return safe_name[:120]
